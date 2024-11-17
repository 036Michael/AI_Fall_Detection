import pandas as pd
import time
from openpyxl import Workbook, load_workbook
import random
import os

# 創建一個新的 Excel 文件
file_name = 'data.xlsx'
# 如果文件存在，打開並附加數據；否則，創建一個新的文件
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Timestamp', 'fall_duration'])  # 添加表頭

# 模擬即時接收數據並寫入 Excel
def simulate_and_write_to_excel(data = None):
    if data is None:
        return
    
    # 將數據寫入 Excel
    ws.append([data['timestamp'], data['fall_duration']])
    wb.save(file_name)  # 儲存 Excel 文件
    # print(f"Inserted data: {timestamp}, {nose}")
    
    # time.sleep(5)  # 每 5 秒插入一次數據

# 開始數據接收與寫入
# simulate_and_write_to_excel()
